from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import *
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from selenium.webdriver.chrome.service import Service
import time


url = "https://www.ajio.com/men-jackets-coats/c/830216010"
driver = webdriver.Chrome(service=Service(executable_path=ChromeDriverManager().install()))
driver.get(url)
driver.maximize_window()
last = driver.execute_script("return document.body.scrollHeight")
i = 0
action = ActionChains(driver)

time.sleep(10)
while True:
    print(i)
    el1 = driver.find_element(By.CSS_SELECTOR,"div.dekstop-footer-content")
    jscode = "arguments[0].scrollIntoView();"
    driver.execute_script(jscode, el1)
    # action.send_keys(Keys.PAGE_UP)
    time.sleep(2)
    action.send_keys(Keys.UP)
    time.sleep(5)
    new = driver.execute_script("return document.body.scrollHeight")
    print(new)
    print(last)
    if last == new:
        break
    last = new
    i += 1
    if i == 10:
        break
workbook = openpyxl.Workbook()
sheet = workbook.active

brandName = driver.find_elements(By.CSS_SELECTOR,"div.brand")
productName = driver.find_elements(By.CSS_SELECTOR,"div.nameCls")
priceTag = driver.find_elements(By.CSS_SELECTOR,"span.price")

sheet.cell(row=1, column=1,value="Brand_Name")
for i in range (len(brandName)):
    sheet.cell(row=2+i, column=1, value = brandName[i].text)
    
sheet.cell(row=1, column=2,value="Product_Name")
for j in range (len(productName)):
    sheet.cell(row=2+j, column=2, value = productName[j].text)

sheet.cell(row=1, column=3,value="Price")
for k in range (len(priceTag)):
    price = priceTag[k].text.replace("₹","").replace(",","").strip()
    sheet.cell(row=2+k, column=3, value = int(price))

workbook.save("MyWorkBook.xlsx")