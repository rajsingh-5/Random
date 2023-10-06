# Loading files ending with .xlsx and writting their path in new excel file and again opening each excel file to
# fetch the data at present at 2,2 and writting data value same excel file where path is written
import os

import openpyxl

files = []
directory = "TestCases"
path = os.path.join(os.getcwd(), directory)

p = os.listdir(os.getcwd())
if "TestCases" in p:
    for file in os.listdir('TestCases'):
        if file.endswith(".xlsx"):
            newPath = os.path.join(path, file)
            files.append(newPath)
files.sort()
workbook = openpyxl.Workbook()
workbook.create_sheet("Sheet1")
sh = workbook.active
row = 2

# Write headers
sh.cell(1, 1, "TestCases")
sh.cell(1, 2, "Path")

for i in range(len(files)):
    wb = openpyxl.load_workbook(files[i])
    sheet = wb.active
    value = sheet.cell(2, 2).value
    sh.cell(row + i, 1, value)  # TestCase_Name
    sh.cell(row + i, 2, files[i])  # Path

workbook.save("path.xlsx")
workbook.close()
