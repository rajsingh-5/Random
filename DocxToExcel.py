import openpyxl
import pandas as pd
from docx import Document

workbook = openpyxl.Workbook()
sheet = workbook.active

doc = Document(r"C:\Users\ACER\Downloads\Test Report API_ACKMSG_GLS.docx")

df = pd.DataFrame()
for table in doc.tables:
    for row in table.rows:
        text = [cell.text for cell in row.cells]
        df = df.append([text], ignore_index=True)
lis = list(df[0][7:])
d = [i for i in lis if "TC No" in i]

data = []
header = []
for i in lis:
    s = i.split(":")
    s[0].strip()
    s[1].strip()
    header.append(s[0])
    data.append(s[1])
row = 1
for i in range(0, 5):
    sheet.cell(row=row, column=i + 1, value=header[i])
j = 0
for data_row in range(2, len(d)+2):
    for data_column in range(1, 6):
        d = data[j]
        sheet.cell(row=data_row, column=data_column, value=data[j].replace("Request", "").replace("OLD", "").strip())
        j += 1
lis1 = list(df[1][7:])
status_lis = [item for item in lis1 if "Status:" in item]
j = 0
sheet.cell(row=1, column=3, value="Status")
for data_row in range(2, len(status_lis)+2):
    status_ = status_lis[j].split()
    sheet.cell(row=data_row, column=3, value=status_[1].replace("Request", "").replace("OLD", "").strip())
    j += 1
workbook.save("file.xlsx")
